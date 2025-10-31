# make_json.py
# 251022중등재원생.xlsx → teachers-middle.json / students-by-teacher-middle.json 자동 생성
# 개선 사항:
#  - argparse 지원: 입력/출력/시트/정렬옵션을 커맨드라인에서 지정 가능
#  - 헤더 매칭 강화: 공백/개행/대소문자 변형, 동의어 다중 처리
#  - 데이터 정규화: 문자열 trim, 식별번호/성명/담임 비어있으면 스킵
#  - 중복 처리: 동일 담임 밑 중복 식별번호는 마지막 레코드로 덮어쓰기
#  - 정렬 일관화: 담임명 정렬(대소문자 무시), 담임별 학생목록 정렬(학년→성명→식별번호)
#  - 출력 폴더 자동 생성 및 요약 로그

from openpyxl import load_workbook
import json, os, sys, argparse
from typing import Dict, Any, List, Optional

# -------------------------
# 유틸
# -------------------------
def norm(s) -> str:
    """셀 값 문자열 정규화: None→'', 공백/개행 제거, 좌우 trim"""
    if s is None:
        return ""
    return str(s).replace("\n", " ").replace("\r", " ").strip()

def build_header_map(ws) -> Dict[str, int]:
    """1행 헤더를 읽어 표준화된 키→열번호(dict) 생성"""
    headers: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = norm(ws.cell(row=1, column=c).value)
        if v:
            headers[v.lower()] = c
    return headers

def find_col(headers: Dict[str, int], *names) -> Optional[int]:
    """헤더 맵에서 동의어 리스트 중 첫 매칭 컬럼번호 반환(소문자 비교)"""
    for name in names:
        key = norm(name).lower()
        if key in headers:
            return headers[key]
    return None

def require_cols(col_map: Dict[str, Optional[int]]):
    missing = [name for name, idx in col_map.items() if idx is None]
    if missing:
        raise SystemExit(f"❌ 필수 컬럼을 찾을 수 없습니다: {', '.join(missing)}. 엑셀 헤더를 확인하세요.")

def int_or_str(v):
    """학년 등 수치형이 들어오면 int로, 아니면 원문 문자열"""
    t = norm(v)
    if t == "":
        return None
    try:
        return int(float(t))
    except:
        return t

# -------------------------
# 메인 로직
# -------------------------
def run(src_xlsx: str, out_dir: str, sheet_name: Optional[str], pretty: bool, sort_case_insensitive: bool):
    if not os.path.exists(src_xlsx):
        raise SystemExit(f"❌ 입력 파일을 찾을 수 없습니다: {src_xlsx}")

    os.makedirs(out_dir, exist_ok=True)
    out_teachers = os.path.join(out_dir, "teachers-middle.json")
    out_byteach = os.path.join(out_dir, "students-by-teacher-middle.json")

    wb = load_workbook(src_xlsx, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"❌ 지정한 시트를 찾을 수 없습니다: {sheet_name}\n사용 가능 시트: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb.active

    headers = build_header_map(ws)

    # 컬럼 탐색(동의어 포함)
    col_id   = find_col(headers, "식별번호", "id", "학생식별번호")
    col_name = find_col(headers, "성명", "이름", "학생명")
    col_sch  = find_col(headers, "학교명", "학교")
    col_grade= find_col(headers, "학년", "grade")
    col_class= find_col(headers, "반명", "클래스", "class")
    col_level= find_col(headers, "레벨", "level")
    col_hm   = find_col(headers, "담임", "담임명", "담임선생님", "담임교사", "담임선생님명")

    # 필수 컬럼 검증
    require_cols({
        "식별번호": col_id,
        "성명": col_name,
        "담임": col_hm,
    })

    # 담임별 학생 모으기(중복 식별번호 마지막 값 우선)
    # 구조: { "담임명": { "식별번호": rec, ... } }
    grouped: Dict[str, Dict[str, Dict[str, Any]]] = {}

    rows_total = 0
    rows_kept  = 0
    for r in range(2, ws.max_row + 1):
        rows_total += 1
        sid = norm(ws.cell(row=r, column=col_id).value)
        name = norm(ws.cell(row=r, column=col_name).value)
        teacher = norm(ws.cell(row=r, column=col_hm).value)

        # 필수값 검증
        if not sid or not name or not teacher:
            continue

        rec = {
            "식별번호": sid,
            "성명": name,
            "학교명": norm(ws.cell(row=r, column=col_sch).value) if col_sch else None,
            "학년": int_or_str(ws.cell(row=r, column=col_grade).value) if col_grade else None,
            "레벨": norm(ws.cell(row=r, column=col_level).value) if col_level else None,
            "반명": norm(ws.cell(row=r, column=col_class).value) if col_class else None
        }

        # 담임 그룹 초기화
        if teacher not in grouped:
            grouped[teacher] = {}
        # 같은 담임 아래 동일 식별번호가 여러번 나오면 마지막 레코드로 덮어쓰기
        grouped[teacher][sid] = rec
        rows_kept += 1

    # 담임명 정렬
    teachers: List[str] = list(grouped.keys())
    if sort_case_insensitive:
        teachers.sort(key=lambda s: s.lower())
    else:
        teachers.sort()

    # 담임별 리스트로 변환 + 학생 정렬(학년→성명→식별번호)
    by_teacher_list: Dict[str, List[Dict[str, Any]]] = {}
    for t in teachers:
        students = list(grouped[t].values())
        students.sort(
            key=lambda x: (
                (x.get("학년") if isinstance(x.get("학년"), int) else 9999),
                str(x.get("성명") or ""),
                str(x.get("식별번호") or "")
            )
        )
        by_teacher_list[t] = students

    # JSON 저장
    json_kwargs = {"ensure_ascii": False, "indent": 2} if pretty else {"ensure_ascii": False}
    with open(out_teachers, "w", encoding="utf-8") as f:
        json.dump(teachers, f, **json_kwargs)
    with open(out_byteach, "w", encoding="utf-8") as f:
        json.dump(by_teacher_list, f, **json_kwargs)

    # 요약 출력
    teacher_count = len(teachers)
    student_count = sum(len(v) for v in by_teacher_list.values())
    print("✅ 생성 완료")
    print(f" - 입력: {src_xlsx} (시트: {ws.title})")
    print(f" - 총 행수: {rows_total:,} / 사용 행수(필수값 만족): {rows_kept:,}")
    print(f" - 담임 수: {teacher_count:,} / 학생 수(중복 제거 후): {student_count:,}")
    print(f" - 출력(담임목록): {out_teachers}")
    print(f" - 출력(담임별학생): {out_byteach}")

# -------------------------
# CLI
# -------------------------
def main():
    parser = argparse.ArgumentParser(description="251022중등재원생.xlsx → teachers-middle.json / students-by-teacher-middle.json 생성")
    parser.add_argument("--src", default=r"C:\dev\samboh-web\251022중등재원생.xlsx",
                        help="입력 엑셀 경로 (기본: C:\\dev\\samboh-web\\251022중등재원생.xlsx)")
    parser.add_argument("--out", default=r"C:\dev\samboh-web\public\data",
                        help="출력 폴더 경로 (기본: C:\\dev\\samboh-web\\public\\data)")
    parser.add_argument("--sheet", default=None, help="시트 이름(미지정 시 active 시트)")
    parser.add_argument("--pretty", action="store_true", help="JSON 들여쓰기(가독성용)")
    parser.add_argument("--case-insensitive-sort", action="store_true", help="담임명 정렬 시 대소문자 무시")
    args = parser.parse_args()

    run(
        src_xlsx=args.src,
        out_dir=args.out,
        sheet_name=args.sheet,
        pretty=args.pretty,
        sort_case_insensitive=args.case_insensitive_sort
    )

if __name__ == "__main__":
    main()
